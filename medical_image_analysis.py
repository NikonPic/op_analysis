# %%
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Medical Imaging Data Analysis Script - Analysis Module
Performs comprehensive comparison between AI and orthopedic surgeon measurements
"""

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
from typing import Dict, List, Tuple, Set, Any, Optional, Union
import logging
from pathlib import Path
import statsmodels.api as sm
from itertools import combinations
from openpyxl.utils import get_column_letter

# Import the data loading functions from your previous script
# Assuming the previous script is saved as data_loader.py
from prepare_data import load_surgeon_data, load_ai_data, SURGEON_SHEET_NAMES, OP_MODES, RELEVANT_ANGLES, RELEVANT_ANGLES_PREOP, RELEVANT_ANGLES_POSTOP

# Configure logging
logging.basicConfig(
    level=logging.INFO, 
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Create output directories
OUTPUT_DIR = Path("./analysis_results")
OUTPUT_DIR.mkdir(exist_ok=True)
(OUTPUT_DIR / "figures").mkdir(exist_ok=True)
(OUTPUT_DIR / "tables").mkdir(exist_ok=True)
(OUTPUT_DIR / "stats").mkdir(exist_ok=True)

# Analysis settings
BLAND_ALTMAN_CONFIDENCE = 0.95  # 95% confidence interval
CORRELATION_METHODS = ["pearson", "spearman"]
ERROR_METRICS = ["MAE", "RMSE", "ME", "Median_AE"]  # Mean Absolute Error, Root Mean Square Error, Mean Error, Median Absolute Error

# Set seaborn style for better visualization
sns.set_theme(style="whitegrid")
sns.set_palette("colorblind")
plt.rcParams.update({'font.size': 12})

def calculate_mean_os_data(os_data: Dict) -> Dict:
    """
    Calculate mean values across all surgeons for each measurement.
    Handles missing, NaN, and invalid values by excluding them from mean calculation.
   
    Args:
        os_data (Dict): Orthopedic surgeon data
       
    Returns:
        Dict: Mean orthopedic surgeon data
    """
    mean_os_data = {}
    logger.info("Calculating mean orthopedic surgeon data")
   
    for dataset_mode in ["int", "ex"]:
        mean_os_data[dataset_mode] = {}
       
        for opmode in OP_MODES:
            mean_os_data[dataset_mode][opmode] = {}
           
            for angle in RELEVANT_ANGLES:
                mean_os_data[dataset_mode][opmode][angle] = {}
               
                try:
                    # Collect all patient_ids across all surgeons for this combination
                    all_patient_ids = set()
                    for sheet_name in SURGEON_SHEET_NAMES:
                        try:
                            if (sheet_name in os_data and 
                                dataset_mode in os_data[sheet_name] and 
                                opmode in os_data[sheet_name][dataset_mode] and 
                                angle in os_data[sheet_name][dataset_mode][opmode]):
                                all_patient_ids.update(os_data[sheet_name][dataset_mode][opmode][angle].keys())
                        except Exception as e:
                            logger.warning(f"Error collecting patient IDs for {sheet_name}, {dataset_mode}, {opmode}, {angle}: {e}")
                            continue
                    
                    if not all_patient_ids:
                        logger.info(f"No patient IDs found for {dataset_mode}, {opmode}, {angle}")
                        continue
                   
                    # For each patient, calculate mean across surgeons
                    for patient_id in all_patient_ids:
                        valid_values = []
                        for sheet_name in SURGEON_SHEET_NAMES:
                            try:
                                if (sheet_name in os_data and 
                                    dataset_mode in os_data[sheet_name] and 
                                    opmode in os_data[sheet_name][dataset_mode] and 
                                    angle in os_data[sheet_name][dataset_mode][opmode]):
                                    
                                    value = os_data[sheet_name][dataset_mode][opmode][angle].get(patient_id)
                                    
                                    # Only include valid numerical values (not None, not NaN)
                                    if value is not None and not pd.isna(value):
                                        # Check if value is numerical
                                        try:
                                            float_value = float(value)
                                            valid_values.append(float_value)
                                        except (ValueError, TypeError):
                                            logger.warning(f"Non-numerical value for {patient_id}, {sheet_name}, {angle}: {value}")
                            except Exception as e:
                                logger.warning(f"Error retrieving value for {patient_id}, {sheet_name}, {dataset_mode}, {opmode}, {angle}: {e}")
                                continue
                       
                        # Only calculate mean if we have at least one valid value
                        if valid_values:
                            mean_os_data[dataset_mode][opmode][angle][patient_id] = np.mean(valid_values)
                            
                        if len(valid_values) < len(SURGEON_SHEET_NAMES):
                            logger.info(f"Patient {patient_id}, {angle}: Only {len(valid_values)} out of {len(SURGEON_SHEET_NAMES)} surgeons have valid measurements")
                
                except Exception as e:
                    logger.error(f"Error calculating mean for {dataset_mode}, {opmode}, {angle}: {e}")
    
    return mean_os_data

def extract_common_measurements(data1: Dict, data2: Dict, 
                               dataset_mode: str, opmode: str, angle: str) -> Tuple[List, List]:
    """
    Extract paired measurements where both sources have data.
    
    Args:
        data1 (Dict): First data source
        data2 (Dict): Second data source
        dataset_mode (str): Dataset mode (int/ex)
        opmode (str): Operation mode
        angle (str): Angle measurement
        
    Returns:
        Tuple[List, List]: Paired measurements from both sources
    """
    values1 = []
    values2 = []
    
    try:
        data1_patients = data1[dataset_mode][opmode][angle]
        data2_patients = data2[dataset_mode][opmode][angle]
        
        # Find common patient_ids
        common_patients = set(data1_patients.keys()) & set(data2_patients.keys())
        
        for patient_id in common_patients:
            values1.append(data1_patients[patient_id])
            values2.append(data2_patients[patient_id])
            
        return values1, values2
    
    except KeyError as e:
        logger.warning(f"Key error when extracting common measurements: {e}")
        return [], []

def calculate_error_metrics(actual: List[float], predicted: List[float]) -> Dict[str, float]:
    """
    Calculate error metrics between two sets of measurements.
    
    Args:
        actual (List[float]): Actual values (ground truth)
        predicted (List[float]): Predicted values
        
    Returns:
        Dict[str, float]: Dictionary of error metrics
    """
    if not actual or not predicted or len(actual) != len(predicted):
        return {metric: np.nan for metric in ERROR_METRICS}
    
    actual = np.array(actual)
    predicted = np.array(predicted)
    
    # Mean Error (ME) - indicates bias direction
    me = np.mean(predicted - actual)
    
    # Mean Absolute Error (MAE)
    mae = np.mean(np.abs(predicted - actual))
    
    # Root Mean Square Error (RMSE)
    rmse = np.sqrt(np.mean((predicted - actual) ** 2))
    
    return {
        "ME": me,
        "MAE": mae,
        "RMSE": rmse
    }

def calculate_correlation_metrics(x: List[float], y: List[float]) -> Dict[str, float]:
    """
    Calculate correlation metrics between two sets of measurements.
    
    Args:
        x (List[float]): First set of measurements
        y (List[float]): Second set of measurements
        
    Returns:
        Dict[str, float]: Dictionary of correlation metrics
    """
    if not x or not y or len(x) != len(y):
        return {method: np.nan for method in CORRELATION_METHODS}
    
    results = {}
    
    # Pearson correlation
    pearson_r, pearson_p = stats.pearsonr(x, y)
    results["pearson"] = pearson_r
    results["pearson_p"] = pearson_p
    
    # Spearman correlation
    spearman_r, spearman_p = stats.spearmanr(x, y)
    results["spearman"] = spearman_r
    results["spearman_p"] = spearman_p
    
    return results

def calculate_icc(x: List[float], y: List[float]) -> float:
    """
    Calculate Intraclass Correlation Coefficient (ICC).
    
    Args:
        x (List[float]): First set of measurements
        y (List[float]): Second set of measurements
        
    Returns:
        float: ICC value
    """
    if not x or not y or len(x) != len(y) or len(x) < 2:
        return np.nan
    
    try:
        # Create a dataframe with data in long format
        data = []
        for i in range(len(x)):
            data.append({"subject": i, "rater": 1, "measurement": x[i]})
            data.append({"subject": i, "rater": 2, "measurement": y[i]})
        
        df = pd.DataFrame(data)
        
        # Fit a mixed effects model
        formula = "measurement ~ 1"
        model = sm.MixedLM.from_formula(formula, groups="subject", data=df)
        result = model.fit()
        
        # Calculate ICC based on the model
        subj_var = result.cov_re.iloc[0, 0]
        resid_var = result.scale
        icc = subj_var / (subj_var + resid_var)
        
        return icc
    
    except Exception as e:
        logger.warning(f"Error calculating ICC: {e}")
        return np.nan

def create_bland_altman_plot(x: List[float], y: List[float], 
                           title: str = "", 
                           save_path: Optional[str] = None) -> plt.Figure:
    """
    Create a Bland-Altman plot to assess agreement between two methods.
    
    Args:
        x (List[float]): Measurements from method 1
        y (List[float]): Measurements from method 2
        title (str): Plot title
        save_path (str, optional): Path to save the figure
        
    Returns:
        plt.Figure: The created figure
    """
    if not x or not y or len(x) != len(y):
        logger.warning("Cannot create Bland-Altman plot with empty or unequal arrays")
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.text(0.5, 0.5, "Insufficient data for Bland-Altman plot", 
                horizontalalignment='center', verticalalignment='center')
        return fig
    
    # Convert to numpy arrays
    x = np.array(x)
    y = np.array(y)
    
    # Calculate mean and difference
    mean = np.mean([x, y], axis=0)
    diff = y - x  # Method 2 - Method 1
    
    # Calculate mean difference and standard deviation
    md = np.mean(diff)
    sd = np.std(diff, axis=0)
    
    # Calculate limits of agreement
    upper_loa = md + 1.96 * sd
    lower_loa = md - 1.96 * sd
    
    # Create the plot
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Plot scatter points
    ax.scatter(mean, diff, alpha=0.7)
    
    # Add horizontal lines for mean difference and limits of agreement
    ax.axhline(md, color='k', linestyle='-', linewidth=1)
    ax.axhline(upper_loa, color='r', linestyle='--', linewidth=1)
    ax.axhline(lower_loa, color='r', linestyle='--', linewidth=1)
    
    # Add text annotations
    ax.text(np.max(mean), md, f'Mean diff: {md:.2f}', 
            horizontalalignment='right', verticalalignment='bottom')
    ax.text(np.max(mean), upper_loa, f'Upper LoA: {upper_loa:.2f}', 
            horizontalalignment='right', verticalalignment='bottom')
    ax.text(np.max(mean), lower_loa, f'Lower LoA: {lower_loa:.2f}', 
            horizontalalignment='right', verticalalignment='top')
    
    # Set labels and title
    ax.set_xlabel('Mean of Methods')
    ax.set_ylabel('Difference between Methods (Method 2 - Method 1)')
    ax.set_title(title)
    
    # Add grid
    ax.grid(True, linestyle='--', alpha=0.7)
    
    # Tight layout
    plt.tight_layout()
    
    # Save if requested
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
    
    return fig

def create_scatter_plot(x: List[float], y: List[float], 
                      title: str = "", 
                      xlabel: str = "Method 1", 
                      ylabel: str = "Method 2",
                      save_path: Optional[str] = None) -> plt.Figure:
    """
    Create a scatter plot with correlation metrics.
    
    Args:
        x (List[float]): Measurements from method 1
        y (List[float]): Measurements from method 2
        title (str): Plot title
        xlabel (str): X-axis label
        ylabel (str): Y-axis label
        save_path (str, optional): Path to save the figure
        
    Returns:
        plt.Figure: The created figure
    """
    if not x or not y or len(x) != len(y):
        logger.warning("Cannot create scatter plot with empty or unequal arrays")
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.text(0.5, 0.5, "Insufficient data for scatter plot", 
                horizontalalignment='center', verticalalignment='center')
        return fig
    
    # Calculate correlation metrics
    corr_metrics = calculate_correlation_metrics(x, y)
    pearson_r = corr_metrics.get("pearson", np.nan)
    
    # Create the plot
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Plot scatter points
    ax.scatter(x, y, alpha=0.7)
    
    # Add regression line
    if not np.isnan(pearson_r) and len(x) > 1:
        m, b = np.polyfit(x, y, 1)
        x_line = np.linspace(min(x), max(x), 100)
        y_line = m * x_line + b
        ax.plot(x_line, y_line, 'r-', linewidth=1)
    
    # Add identity line (y=x)
    ax.plot([min(x), max(x)], [min(x), max(x)], 'k--', alpha=0.5)
    
    # Add correlation info
    if not np.isnan(pearson_r):
        text = f'Pearson r: {pearson_r:.3f}\n'
        text += f'Spearman r: {corr_metrics.get("spearman", np.nan):.3f}'
        ax.text(0.05, 0.95, text, transform=ax.transAxes, 
                verticalalignment='top', horizontalalignment='left',
                bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
    
    # Set labels and title
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    
    # Add grid
    ax.grid(True, linestyle='--', alpha=0.7)
    
    # Set equal aspect ratio
    ax.set_aspect('equal', adjustable='box')
    
    # Tight layout
    plt.tight_layout()
    
    # Save if requested
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
    
    return fig

def analyze_mean_os_vs_ai(mean_os_data: Dict, ai_data: Dict) -> Dict:
    """
    Analyze agreement between mean OS and AI measurements.
    
    Args:
        mean_os_data (Dict): Mean orthopedic surgeon data
        ai_data (Dict): AI data
        
    Returns:
        Dict: Analysis results
    """
    results = {}
    
    for dataset_mode in ["int", "ex"]:
        results[dataset_mode] = {}
        
        for op_id, opmode in enumerate(OP_MODES):
            print(opmode)
            results[dataset_mode][opmode] = {}
            
            cur_rel_angles = RELEVANT_ANGLES_POSTOP
            if op_id == 0:
                cur_rel_angles = RELEVANT_ANGLES_PREOP
            
            for angle in RELEVANT_ANGLES:
                # Extract common measurements
                os_values, ai_values = extract_common_measurements(
                    mean_os_data, ai_data, dataset_mode, opmode, angle
                )
                
                if not os_values or not ai_values:
                    results[dataset_mode][opmode][angle] = {
                        "n_samples": 0,
                        "error_metrics": {metric: np.nan for metric in ERROR_METRICS},
                        "correlation": {method: np.nan for method in CORRELATION_METHODS},
                        "icc": np.nan
                    }
                    continue
                
                # Calculate metrics
                error_metrics = calculate_error_metrics(os_values, ai_values)
                correlation = calculate_correlation_metrics(os_values, ai_values)
                icc = calculate_icc(os_values, ai_values)
                
                # Store results
                results[dataset_mode][opmode][angle] = {
                    "n_samples": len(os_values),
                    "error_metrics": error_metrics,
                    "correlation": correlation,
                    "icc": icc
                }
                
                # Create and save plots if we have enough data
                if len(os_values) >= 3:
                    # Directory for this specific comparison
                    plot_dir = OUTPUT_DIR / "figures" / f"{dataset_mode}_{opmode}" / "mean_os_vs_ai"
                    plot_dir.mkdir(parents=True, exist_ok=True)
                    
                    # Bland-Altman plot
                    title = f"Bland-Altman: Mean OS vs AI - {angle} ({dataset_mode}, {opmode})"
                    save_path = plot_dir / f"bland_altman_{angle}.png"
                    create_bland_altman_plot(
                        os_values, ai_values, title=title, save_path=str(save_path)
                    )
                    
                    # Scatter plot
                    title = f"Correlation: Mean OS vs AI - {angle} ({dataset_mode}, {opmode})"
                    save_path = plot_dir / f"scatter_{angle}.png"
                    create_scatter_plot(
                        os_values, ai_values, title=title, 
                        xlabel="Mean OS Measurement", ylabel="AI Measurement",
                        save_path=str(save_path)
                    )
    
    return results

def analyze_individual_os_vs_ai(os_data: Dict, ai_data: Dict) -> Dict:
    """
    Analyze agreement between individual OS and AI measurements.
    
    Args:
        os_data (Dict): Orthopedic surgeon data
        ai_data (Dict): AI data
        
    Returns:
        Dict: Analysis results
    """
    results = {}
    
    for sheet_name in SURGEON_SHEET_NAMES:
        results[sheet_name] = {}
        
        for dataset_mode in ["int", "ex"]:
            results[sheet_name][dataset_mode] = {}
            
            for opmode in OP_MODES:
                results[sheet_name][dataset_mode][opmode] = {}
                
                for angle in RELEVANT_ANGLES:
                    # Extract the surgeon's data for this combination
                    try:
                        if sheet_name not in os_data or \
                           dataset_mode not in os_data[sheet_name] or \
                           opmode not in os_data[sheet_name][dataset_mode] or \
                           angle not in os_data[sheet_name][dataset_mode][opmode]:
                            # Skip if data is missing at any level
                            continue
                        
                        surgeon_data = {
                            dataset_mode: {
                                opmode: {
                                    angle: os_data[sheet_name][dataset_mode][opmode][angle]
                                }
                            }
                        }
                        
                        # Extract common measurements
                        os_values, ai_values = extract_common_measurements(
                            surgeon_data, ai_data, dataset_mode, opmode, angle
                        )
                        
                        if not os_values or not ai_values:
                            results[sheet_name][dataset_mode][opmode][angle] = {
                                "n_samples": 0,
                                "error_metrics": {metric: np.nan for metric in ERROR_METRICS},
                                "correlation": {method: np.nan for method in CORRELATION_METHODS},
                                "icc": np.nan
                            }
                            continue
                        
                        # Calculate metrics
                        error_metrics = calculate_error_metrics(os_values, ai_values)
                        correlation = calculate_correlation_metrics(os_values, ai_values)
                        icc = calculate_icc(os_values, ai_values)
                        
                        # Store results
                        results[sheet_name][dataset_mode][opmode][angle] = {
                            "n_samples": len(os_values),
                            "error_metrics": error_metrics,
                            "correlation": correlation,
                            "icc": icc
                        }
                        
                        # Create and save plots if we have enough data
                        if len(os_values) >= 3:
                            # Directory for this specific comparison
                            plot_dir = OUTPUT_DIR / "figures" / f"{dataset_mode}_{opmode}" / f"{sheet_name}_vs_ai"
                            plot_dir.mkdir(parents=True, exist_ok=True)
                            
                            # Bland-Altman plot
                            title = f"Bland-Altman: {sheet_name} vs AI - {angle} ({dataset_mode}, {opmode})"
                            save_path = plot_dir / f"bland_altman_{angle}.png"
                            create_bland_altman_plot(
                                os_values, ai_values, title=title, save_path=str(save_path)
                            )
                            
                            # Scatter plot
                            title = f"Correlation: {sheet_name} vs AI - {angle} ({dataset_mode}, {opmode})"
                            save_path = plot_dir / f"scatter_{angle}.png"
                            create_scatter_plot(
                                os_values, ai_values, title=title, 
                                xlabel=f"{sheet_name} Measurement", ylabel="AI Measurement",
                                save_path=str(save_path)
                            )
                    
                    except Exception as e:
                        logger.error(f"Error in individual OS vs AI analysis for {sheet_name}, "
                                    f"{dataset_mode}, {opmode}, {angle}: {e}")
    
    return results

def analyze_os_vs_os(os_data: Dict) -> Dict:
    """
    Analyze inter-observer agreement between orthopedic surgeons.
    
    Args:
        os_data (Dict): Orthopedic surgeon data
        
    Returns:
        Dict: Analysis results
    """
    results = {}
    
    # Get all pairs of surgeons
    surgeon_pairs = list(combinations(SURGEON_SHEET_NAMES, 2))
    
    for dataset_mode in ["int", "ex"]:
        results[dataset_mode] = {}
        
        for opmode in OP_MODES:
            results[dataset_mode][opmode] = {}
            
            for angle in RELEVANT_ANGLES:
                results[dataset_mode][opmode][angle] = {}
                
                # Analyze each pair of surgeons
                for surgeon1, surgeon2 in surgeon_pairs:
                    pair_name = f"{surgeon1}_vs_{surgeon2}"
                    
                    # Extract individual surgeon data for this combination
                    try:
                        surgeon1_data = {
                            dataset_mode: {
                                opmode: {
                                    angle: os_data[surgeon1][dataset_mode][opmode][angle]
                                }
                            }
                        }
                        
                        surgeon2_data = {
                            dataset_mode: {
                                opmode: {
                                    angle: os_data[surgeon2][dataset_mode][opmode][angle]
                                }
                            }
                        }
                        
                        # Extract common measurements
                        surgeon1_values, surgeon2_values = extract_common_measurements(
                            surgeon1_data, surgeon2_data, dataset_mode, opmode, angle
                        )
                        
                        if not surgeon1_values or not surgeon2_values:
                            results[dataset_mode][opmode][angle][pair_name] = {
                                "n_samples": 0,
                                "error_metrics": {metric: np.nan for metric in ERROR_METRICS},
                                "correlation": {method: np.nan for method in CORRELATION_METHODS},
                                "icc": np.nan
                            }
                            continue
                        
                        # Calculate metrics
                        error_metrics = calculate_error_metrics(surgeon1_values, surgeon2_values)
                        correlation = calculate_correlation_metrics(surgeon1_values, surgeon2_values)
                        icc = calculate_icc(surgeon1_values, surgeon2_values)
                        
                        # Store results
                        results[dataset_mode][opmode][angle][pair_name] = {
                            "n_samples": len(surgeon1_values),
                            "error_metrics": error_metrics,
                            "correlation": correlation,
                            "icc": icc
                        }
                        
                        # Create and save plots if we have enough data
                        if len(surgeon1_values) >= 3:
                            # Directory for this specific comparison
                            plot_dir = OUTPUT_DIR / "figures" / f"{dataset_mode}_{opmode}" / "os_vs_os"
                            plot_dir.mkdir(parents=True, exist_ok=True)
                            
                            # Bland-Altman plot
                            title = f"Bland-Altman: {surgeon1} vs {surgeon2} - {angle} ({dataset_mode}, {opmode})"
                            save_path = plot_dir / f"bland_altman_{pair_name}_{angle}.png"
                            create_bland_altman_plot(
                                surgeon1_values, surgeon2_values, title=title, save_path=str(save_path)
                            )
                            
                            # Scatter plot
                            title = f"Correlation: {surgeon1} vs {surgeon2} - {angle} ({dataset_mode}, {opmode})"
                            save_path = plot_dir / f"scatter_{pair_name}_{angle}.png"
                            create_scatter_plot(
                                surgeon1_values, surgeon2_values, title=title, 
                                xlabel=f"{surgeon1} Measurement", ylabel=f"{surgeon2} Measurement",
                                save_path=str(save_path)
                            )
                    
                    except Exception as e:
                        logger.error(f"Error in OS vs OS analysis for {pair_name}, "
                                    f"{dataset_mode}, {opmode}, {angle}: {e}")
    
    return results

def export_results_to_csv(results: Dict, filename: str) -> None:
    """
    Export analysis results to CSV file.
    
    Args:
        results (Dict): Analysis results
        filename (str): Output filename
    """
    output_path = OUTPUT_DIR / "tables" / filename
    
    # Flatten nested dictionary into rows for CSV
    rows = []
    
    # Handle different result structures
    if "int" in results.keys():  # Mean OS vs AI or OS vs OS structure
        for dataset_mode in ["int", "ex"]:
            for opmode in OP_MODES:
                for angle in RELEVANT_ANGLES:
                    if dataset_mode in results and opmode in results[dataset_mode] and angle in results[dataset_mode][opmode]:
                        data = results[dataset_mode][opmode][angle]
                        
                        # Handle OS vs OS structure (has pairs)
                        if isinstance(data, dict) and (any(key.endswith("_vs_") for key in data.keys()) or any("_vs_" in key for key in data.keys())):
                            for pair_name, pair_data in data.items():
                                row = {
                                    "Dataset": dataset_mode,
                                    "Operation Mode": opmode,
                                    "Angle": angle,
                                    "Comparison": pair_name,
                                    "n_samples": pair_data.get("n_samples", 0)
                                }
                                
                                # Add error metrics
                                for metric, value in pair_data.get("error_metrics", {}).items():
                                    row[metric] = value
                                
                                # Add correlation metrics
                                for method, value in pair_data.get("correlation", {}).items():
                                    row[f"{method}_r"] = value
                                
                                # Add ICC
                                row["ICC"] = pair_data.get("icc", np.nan)
                                
                                rows.append(row)
                        else:
                            # Standard structure (mean OS vs AI)
                            row = {
                                "Dataset": dataset_mode,
                                "Operation Mode": opmode,
                                "Angle": angle,
                                "n_samples": data.get("n_samples", 0)
                            }
                            
                            # Add error metrics
                            for metric, value in data.get("error_metrics", {}).items():
                                row[metric] = value
                            
                            # Add correlation metrics
                            for method, value in data.get("correlation", {}).items():
                                row[f"{method}_r"] = value
                            
                            # Add ICC
                            row["ICC"] = data.get("icc", np.nan)
                            
                            rows.append(row)
    else:
        for sheet_name in results.keys():  # Individual OS vs AI structure
            for dataset_mode in ["int", "ex"]:
                for opmode in OP_MODES:
                    for angle in RELEVANT_ANGLES:
                        if (sheet_name in results and 
                            dataset_mode in results[sheet_name] and 
                            opmode in results[sheet_name][dataset_mode] and 
                            angle in results[sheet_name][dataset_mode][opmode]):
                            
                            data = results[sheet_name][dataset_mode][opmode][angle]
                            
                            row = {
                                "Surgeon": sheet_name,
                                "Dataset": dataset_mode,
                                "Operation Mode": opmode,
                                "Angle": angle,
                                "n_samples": data.get("n_samples", 0)
                            }
                            
                            # Add error metrics
                            for metric, value in data.get("error_metrics", {}).items():
                                row[metric] = value
                            
                            # Add correlation metrics
                            for method, value in data.get("correlation", {}).items():
                                row[f"{method}_r"] = value
                            
                            # Add ICC
                            row["ICC"] = data.get("icc", np.nan)
                            
                            rows.append(row)
    
    # Create DataFrame and save to CSV
    if rows:
        df = pd.DataFrame(rows)
        df.to_csv(output_path, index=False, sep=';', decimal=',', encoding='utf-8-sig')
        logger.info(f"Results exported to {output_path}")
    else:
        logger.warning(f"No results to export to {output_path}")
                            
                            
def create_summary_plots(mean_os_vs_ai_results: Dict) -> None:
    """
    Create summary plots for the entire analysis.
    
    Args:
        mean_os_vs_ai_results (Dict): Results from mean OS vs AI analysis
    """
    # Create ICC summary plot for each dataset mode
    for dataset_mode in ["int", "ex"]:
        # Prepare data for plotting
        angles = []
        opmodes = []
        icc_values = []
        
        for opmode in OP_MODES:
            for angle in RELEVANT_ANGLES:
                if (dataset_mode in mean_os_vs_ai_results and 
                    opmode in mean_os_vs_ai_results[dataset_mode] and 
                    angle in mean_os_vs_ai_results[dataset_mode][opmode]):
                    
                    data = mean_os_vs_ai_results[dataset_mode][opmode][angle]
                    icc = data.get("icc", np.nan)
                    
                    if not np.isnan(icc):
                        angles.append(angle)
                        opmodes.append(opmode)
                        icc_values.append(icc)
        
        if not angles:
            continue  # Skip if no data
        
        # Create DataFrame for plotting
        df = pd.DataFrame({
            "Angle": angles,
            "Operation Mode": opmodes,
            "ICC": icc_values
        })
        
        # Create directory for summary plots
        summary_dir = OUTPUT_DIR / "figures" / "summary"
        summary_dir.mkdir(parents=True, exist_ok=True)
        
        # Create ICC heatmap
        plt.figure(figsize=(12, 8))
        pivot_table = df.pivot_table(values="ICC", index="Angle", columns="Operation Mode")
        sns.heatmap(pivot_table, annot=True, cmap="YlGnBu", vmin=0, vmax=1, fmt=".2f")
        plt.title(f"ICC: Mean OS vs AI - {dataset_mode.upper()} Dataset")
        plt.tight_layout()
        plt.savefig(summary_dir / f"icc_heatmap_{dataset_mode}.png", dpi=300, bbox_inches='tight')
        plt.close()
        
        # Create ICC barplot
        plt.figure(figsize=(12, 8))
        sns.barplot(x="Angle", y="ICC", hue="Operation Mode", data=df)
        plt.title(f"ICC: Mean OS vs AI - {dataset_mode.upper()} Dataset")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        plt.savefig(summary_dir / f"icc_barplot_{dataset_mode}.png", dpi=300, bbox_inches='tight')
        plt.close()


def create_summary_tables(mean_os_vs_ai_results: Dict, individual_os_vs_ai_results: Dict, os_vs_os_results: Dict) -> None:
    """
    Create summary tables of the analysis results.
    
    Args:
        mean_os_vs_ai_results (Dict): Results from mean OS vs AI analysis
        individual_os_vs_ai_results (Dict): Results from individual OS vs AI analysis
        os_vs_os_results (Dict): Results from OS vs OS analysis
    """
    tables_dir = OUTPUT_DIR / "tables"
    tables_dir.mkdir(exist_ok=True)
    
    # Create ICC summary table for mean OS vs AI
    icc_summary = []
    for dataset_mode in ["int", "ex"]:
        for opmode in OP_MODES:
            for angle in RELEVANT_ANGLES:
                if (dataset_mode in mean_os_vs_ai_results and 
                    opmode in mean_os_vs_ai_results[dataset_mode] and 
                    angle in mean_os_vs_ai_results[dataset_mode][opmode]):
                    
                    data = mean_os_vs_ai_results[dataset_mode][opmode][angle]
                    
                    row = {
                        "Dataset": dataset_mode,
                        "Operation Mode": opmode,
                        "Angle": angle,
                        "ICC": data.get("icc", np.nan),
                        "Alt ICC": data.get("alt_icc", np.nan),
                        "Pearson r": data.get("correlation", {}).get("pearson", np.nan),
                        "Spearman r": data.get("correlation", {}).get("spearman", np.nan),
                        "MAE": data.get("error_metrics", {}).get("MAE", np.nan),
                        "RMSE": data.get("error_metrics", {}).get("RMSE", np.nan),
                        "Median AE": data.get("error_metrics", {}).get("Median_AE", np.nan),
                        "Within 1°": data.get("error_metrics", {}).get("Within_1deg", np.nan),
                        "Within 2°": data.get("error_metrics", {}).get("Within_2deg", np.nan),
                        "Within 3°": data.get("error_metrics", {}).get("Within_3deg", np.nan),
                        "Max Error": data.get("error_metrics", {}).get("Max_Error", np.nan),
                        "t-test p": data.get("statistical_tests", {}).get("t_test_p", np.nan),
                        "Wilcoxon p": data.get("statistical_tests", {}).get("wilcoxon_p", np.nan),
                        "Samples": data.get("n_samples", 0)
                    }
                    
                    icc_summary.append(row)
    
    if icc_summary:
        df = pd.DataFrame(icc_summary)
        df.to_csv(tables_dir / "icc_summary.csv", index=False, sep=';', decimal=',', encoding='utf-8-sig')
        logger.info(f"ICC summary exported to {tables_dir / 'icc_summary.csv'}")
    
    # Create surgeon vs AI summary table
    surgeon_vs_ai_summary = []
    
    for sheet_name in SURGEON_SHEET_NAMES:
        if sheet_name not in individual_os_vs_ai_results:
            continue
            
        for dataset_mode in ["int", "ex"]:
            for opmode in OP_MODES:
                for angle in RELEVANT_ANGLES:
                    if (dataset_mode in individual_os_vs_ai_results[sheet_name] and 
                        opmode in individual_os_vs_ai_results[sheet_name][dataset_mode] and 
                        angle in individual_os_vs_ai_results[sheet_name][dataset_mode][opmode]):
                        
                        data = individual_os_vs_ai_results[sheet_name][dataset_mode][opmode][angle]
                        
                        row = {
                            "Surgeon": sheet_name,
                            "Dataset": dataset_mode,
                            "Operation Mode": opmode,
                            "Angle": angle,
                            "ICC": data.get("icc", np.nan),
                            "MAE": data.get("error_metrics", {}).get("MAE", np.nan),
                            "RMSE": data.get("error_metrics", {}).get("RMSE", np.nan),
                            "Pearson r": data.get("correlation", {}).get("pearson", np.nan),
                            "Samples": data.get("n_samples", 0)
                        }
                        
                        surgeon_vs_ai_summary.append(row)
    
    if surgeon_vs_ai_summary:
        df = pd.DataFrame(surgeon_vs_ai_summary)
        df.to_csv(tables_dir / "surgeon_vs_ai_summary.csv", index=False, sep=';', decimal=',', encoding='utf-8-sig')
        logger.info(f"Surgeon vs AI summary exported to {tables_dir / 'surgeon_vs_ai_summary.csv'}")
    
    # Create inter-surgeon agreement summary
    inter_surgeon_summary = []
    
    for dataset_mode in ["int", "ex"]:
        for opmode in OP_MODES:
            for angle in RELEVANT_ANGLES:
                if (dataset_mode in os_vs_os_results and 
                    opmode in os_vs_os_results[dataset_mode] and 
                    angle in os_vs_os_results[dataset_mode][opmode]):
                    
                    angle_data = os_vs_os_results[dataset_mode][opmode][angle]
                    
                    # Calculate average metrics across all pairs
                    avg_icc = np.nanmean([pair_data.get("icc", np.nan) for pair_data in angle_data.values()])
                    avg_mae = np.nanmean([pair_data.get("error_metrics", {}).get("MAE", np.nan) for pair_data in angle_data.values()])
                    avg_rmse = np.nanmean([pair_data.get("error_metrics", {}).get("RMSE", np.nan) for pair_data in angle_data.values()])
                    avg_pearson = np.nanmean([pair_data.get("correlation", {}).get("pearson", np.nan) for pair_data in angle_data.values()])
                    
                    row = {
                        "Dataset": dataset_mode,
                        "Operation Mode": opmode,
                        "Angle": angle,
                        "Avg ICC": avg_icc,
                        "Avg MAE": avg_mae,
                        "Avg RMSE": avg_rmse,
                        "Avg Pearson r": avg_pearson,
                        "Pairs": len(angle_data)
                    }
                    
                    inter_surgeon_summary.append(row)
    
    if inter_surgeon_summary:
        df = pd.DataFrame(inter_surgeon_summary)
        df.to_csv(tables_dir / "inter_surgeon_summary.csv", index=False, sep=';', decimal=',', encoding='utf-8-sig')
        logger.info(f"Inter-surgeon summary exported to {tables_dir / 'inter_surgeon_summary.csv'}")
        
    # Create combined comparison table (AI vs mean OS vs individual OS)
    combined_comparison = []
    
    for dataset_mode in ["int", "ex"]:
        for opmode in OP_MODES:
            for angle in RELEVANT_ANGLES:
                if (dataset_mode in mean_os_vs_ai_results and 
                    opmode in mean_os_vs_ai_results[dataset_mode] and 
                    angle in mean_os_vs_ai_results[dataset_mode][opmode]):
                    
                    # Get AI vs mean OS data
                    mean_os_data = mean_os_vs_ai_results[dataset_mode][opmode][angle]
                    
                    # Get inter-surgeon agreement data
                    inter_surgeon_icc = np.nan
                    if (dataset_mode in os_vs_os_results and 
                        opmode in os_vs_os_results[dataset_mode] and 
                        angle in os_vs_os_results[dataset_mode][opmode]):
                        
                        pairs_data = os_vs_os_results[dataset_mode][opmode][angle]
                        inter_surgeon_icc = np.nanmean([pair_data.get("icc", np.nan) for pair_data in pairs_data.values()])
                    
                    # Get individual surgeon vs AI data
                    surgeon_iccs = []
                    for sheet_name in SURGEON_SHEET_NAMES:
                        if (sheet_name in individual_os_vs_ai_results and 
                            dataset_mode in individual_os_vs_ai_results[sheet_name] and 
                            opmode in individual_os_vs_ai_results[sheet_name][dataset_mode] and 
                            angle in individual_os_vs_ai_results[sheet_name][dataset_mode][opmode]):
                            
                            icc = individual_os_vs_ai_results[sheet_name][dataset_mode][opmode][angle].get("icc", np.nan)
                            if not np.isnan(icc):
                                surgeon_iccs.append(icc)
                    
                    avg_surgeon_icc = np.nanmean(surgeon_iccs) if surgeon_iccs else np.nan
                    
                    row = {
                        "Dataset": dataset_mode,
                        "Operation Mode": opmode,
                        "Angle": angle,
                        "Mean OS vs AI ICC": mean_os_data.get("icc", np.nan),
                        "Mean OS vs AI MAE": mean_os_data.get("error_metrics", {}).get("MAE", np.nan),
                        "Individual OS vs AI Avg ICC": avg_surgeon_icc,
                        "Inter-surgeon Avg ICC": inter_surgeon_icc,
                        "Samples": mean_os_data.get("n_samples", 0)
                    }
                    
                    combined_comparison.append(row)
    
    if combined_comparison:
        df = pd.DataFrame(combined_comparison)
        df.to_csv(tables_dir / "combined_comparison.csv", index=False, sep=';', decimal=',', encoding='utf-8-sig')
        logger.info(f"Combined comparison exported to {tables_dir / 'combined_comparison.csv'}")

def export_all_metrics_to_excel(mean_os_vs_ai_results: Dict, individual_os_vs_ai_results: Dict, os_vs_os_results: Dict) -> None:
    """
    Export all metrics to a single Excel file with multiple sheets.
    
    Args:
        mean_os_vs_ai_results (Dict): Results from mean OS vs AI analysis
        individual_os_vs_ai_results (Dict): Results from individual OS vs AI analysis
        os_vs_os_results (Dict): Results from OS vs OS analysis
    """
    excel_path = OUTPUT_DIR / "all_metrics.xlsx"
    
    with pd.ExcelWriter(excel_path) as writer:
        # Create Mean OS vs AI sheet
        icc_summary = []
        for dataset_mode in ["int", "ex"]:
            for opmode in OP_MODES:
                for angle in RELEVANT_ANGLES:
                    if (dataset_mode in mean_os_vs_ai_results and 
                        opmode in mean_os_vs_ai_results[dataset_mode] and 
                        angle in mean_os_vs_ai_results[dataset_mode][opmode]):
                        
                        data = mean_os_vs_ai_results[dataset_mode][opmode][angle]
                        
                        row = {
                            "Dataset": dataset_mode,
                            "Operation Mode": opmode,
                            "Angle": angle,
                            "ICC": data.get("icc", np.nan),
                            "Alt ICC": data.get("alt_icc", np.nan),
                            "Pearson r": data.get("correlation", {}).get("pearson", np.nan),
                            "Spearman r": data.get("correlation", {}).get("spearman", np.nan),
                            "MAE": data.get("error_metrics", {}).get("MAE", np.nan),
                            "RMSE": data.get("error_metrics", {}).get("RMSE", np.nan),
                            "Median AE": data.get("error_metrics", {}).get("Median_AE", np.nan),
                            "Within 1°": data.get("error_metrics", {}).get("Within_1deg", np.nan),
                            "Within 2°": data.get("error_metrics", {}).get("Within_2deg", np.nan),
                            "Within 3°": data.get("error_metrics", {}).get("Within_3deg", np.nan),
                            "Max Error": data.get("error_metrics", {}).get("Max_Error", np.nan),
                            "t-test p": data.get("statistical_tests", {}).get("t_test_p", np.nan),
                            "Wilcoxon p": data.get("statistical_tests", {}).get("wilcoxon_p", np.nan),
                            "Samples": data.get("n_samples", 0)
                        }
                        
                        icc_summary.append(row)
        
        if icc_summary:
            df = pd.DataFrame(icc_summary)
            df.to_excel(writer, sheet_name="Mean OS vs AI", index=False)
            logger.info("Mean OS vs AI metrics added to Excel")
        
        # Create Individual OS vs AI sheet
        surgeon_vs_ai_summary = []
        
        for sheet_name in SURGEON_SHEET_NAMES:
            if sheet_name not in individual_os_vs_ai_results:
                continue
                
            for dataset_mode in ["int", "ex"]:
                for opmode in OP_MODES:
                    for angle in RELEVANT_ANGLES:
                        if (dataset_mode in individual_os_vs_ai_results[sheet_name] and 
                            opmode in individual_os_vs_ai_results[sheet_name][dataset_mode] and 
                            angle in individual_os_vs_ai_results[sheet_name][dataset_mode][opmode]):
                            
                            data = individual_os_vs_ai_results[sheet_name][dataset_mode][opmode][angle]
                            
                            row = {
                                "Surgeon": sheet_name,
                                "Dataset": dataset_mode,
                                "Operation Mode": opmode,
                                "Angle": angle,
                                "ICC": data.get("icc", np.nan),
                                "MAE": data.get("error_metrics", {}).get("MAE", np.nan),
                                "RMSE": data.get("error_metrics", {}).get("RMSE", np.nan),
                                "Pearson r": data.get("correlation", {}).get("pearson", np.nan),
                                "Samples": data.get("n_samples", 0)
                            }
                            
                            surgeon_vs_ai_summary.append(row)
        
        if surgeon_vs_ai_summary:
            df = pd.DataFrame(surgeon_vs_ai_summary)
            df.to_excel(writer, sheet_name="Individual OS vs AI", index=False)
            logger.info("Individual OS vs AI metrics added to Excel")
        
        # Create Inter-surgeon Agreement sheet
        inter_surgeon_summary = []
        
        for dataset_mode in ["int", "ex"]:
            for opmode in OP_MODES:
                for angle in RELEVANT_ANGLES:
                    if (dataset_mode in os_vs_os_results and 
                        opmode in os_vs_os_results[dataset_mode] and 
                        angle in os_vs_os_results[dataset_mode][opmode]):
                        
                        angle_data = os_vs_os_results[dataset_mode][opmode][angle]
                        
                        for pair_name, pair_data in angle_data.items():
                            row = {
                                "Surgeon Pair": pair_name,
                                "Dataset": dataset_mode,
                                "Operation Mode": opmode,
                                "Angle": angle,
                                "ICC": pair_data.get("icc", np.nan),
                                "MAE": pair_data.get("error_metrics", {}).get("MAE", np.nan),
                                "RMSE": pair_data.get("error_metrics", {}).get("RMSE", np.nan),
                                "Pearson r": pair_data.get("correlation", {}).get("pearson", np.nan),
                                "Samples": pair_data.get("n_samples", 0)
                            }
                            
                            inter_surgeon_summary.append(row)
        
        if inter_surgeon_summary:
            df = pd.DataFrame(inter_surgeon_summary)
            df.to_excel(writer, sheet_name="Inter-surgeon Agreement", index=False)
            logger.info("Inter-surgeon agreement metrics added to Excel")
        
        # Create Combined Comparison sheet
        combined_comparison = []
        
        for dataset_mode in ["int", "ex"]:
            for opmode in OP_MODES:
                for angle in RELEVANT_ANGLES:
                    if (dataset_mode in mean_os_vs_ai_results and 
                        opmode in mean_os_vs_ai_results[dataset_mode] and 
                        angle in mean_os_vs_ai_results[dataset_mode][opmode]):
                        
                        # Get AI vs mean OS data
                        mean_os_data = mean_os_vs_ai_results[dataset_mode][opmode][angle]
                        
                        # Get inter-surgeon agreement data
                        inter_surgeon_icc = np.nan
                        if (dataset_mode in os_vs_os_results and 
                            opmode in os_vs_os_results[dataset_mode] and 
                            angle in os_vs_os_results[dataset_mode][opmode]):
                            
                            pairs_data = os_vs_os_results[dataset_mode][opmode][angle]
                            inter_surgeon_icc = np.nanmean([pair_data.get("icc", np.nan) for pair_data in pairs_data.values()])
                        
                        # Get individual surgeon vs AI data
                        surgeon_iccs = []
                        for sheet_name in SURGEON_SHEET_NAMES:
                            if (sheet_name in individual_os_vs_ai_results and 
                                dataset_mode in individual_os_vs_ai_results[sheet_name] and 
                                opmode in individual_os_vs_ai_results[sheet_name][dataset_mode] and 
                                angle in individual_os_vs_ai_results[sheet_name][dataset_mode][opmode]):
                                
                                icc = individual_os_vs_ai_results[sheet_name][dataset_mode][opmode][angle].get("icc", np.nan)
                                if not np.isnan(icc):
                                    surgeon_iccs.append(icc)
                        
                        avg_surgeon_icc = np.nanmean(surgeon_iccs) if surgeon_iccs else np.nan
                        
                        row = {
                            "Dataset": dataset_mode,
                            "Operation Mode": opmode,
                            "Angle": angle,
                            "Mean OS vs AI ICC": mean_os_data.get("icc", np.nan),
                            "Mean OS vs AI MAE": mean_os_data.get("error_metrics", {}).get("MAE", np.nan),
                            "Individual OS vs AI Avg ICC": avg_surgeon_icc,
                            "Inter-surgeon Avg ICC": inter_surgeon_icc,
                            "Samples": mean_os_data.get("n_samples", 0)
                        }
                        
                        combined_comparison.append(row)
        
        if combined_comparison:
            df = pd.DataFrame(combined_comparison)
            df.to_excel(writer, sheet_name="Combined Comparison", index=False)
            logger.info("Combined comparison metrics added to Excel")
    
    logger.info(f"All metrics exported to {excel_path}")

def create_comparison_visualizations(mean_os_vs_ai_results: Dict, individual_os_vs_ai_results: Dict, os_vs_os_results: Dict) -> None:
    """
    Create visualizations comparing different agreement metrics.
    
    Args:
        mean_os_vs_ai_results (Dict): Results from mean OS vs AI analysis
        individual_os_vs_ai_results (Dict): Results from individual OS vs AI analysis
        os_vs_os_results (Dict): Results from OS vs OS analysis
    """
    # Create directory for comparison visualizations
    compare_dir = OUTPUT_DIR / "figures" / "comparison"
    compare_dir.mkdir(parents=True, exist_ok=True)
    
    # Extract data for comparison
    comparison_data = []
    
    for dataset_mode in ["int", "ex"]:
        for opmode in OP_MODES:
            for angle in RELEVANT_ANGLES:
                # Get mean OS vs AI data
                if (dataset_mode in mean_os_vs_ai_results and 
                    opmode in mean_os_vs_ai_results[dataset_mode] and 
                    angle in mean_os_vs_ai_results[dataset_mode][opmode]):
                    
                    mean_os_data = mean_os_vs_ai_results[dataset_mode][opmode][angle]
                    mean_os_icc = mean_os_data.get("icc", np.nan)
                    mean_os_mae = mean_os_data.get("error_metrics", {}).get("MAE", np.nan)
                    
                    # Get individual OS vs AI data
                    surgeon_iccs = []
                    surgeon_maes = []
                    
                    for sheet_name in SURGEON_SHEET_NAMES:
                        if (sheet_name in individual_os_vs_ai_results and 
                            dataset_mode in individual_os_vs_ai_results[sheet_name] and 
                            opmode in individual_os_vs_ai_results[sheet_name][dataset_mode] and 
                            angle in individual_os_vs_ai_results[sheet_name][dataset_mode][opmode]):
                            
                            surgeon_data = individual_os_vs_ai_results[sheet_name][dataset_mode][opmode][angle]
                            surgeon_icc = surgeon_data.get("icc", np.nan)
                            surgeon_mae = surgeon_data.get("error_metrics", {}).get("MAE", np.nan)
                            
                            if not np.isnan(surgeon_icc):
                                surgeon_iccs.append(surgeon_icc)
                            
                            if not np.isnan(surgeon_mae):
                                surgeon_maes.append(surgeon_mae)
                    
                    # Get inter-surgeon agreement data
                    inter_surgeon_iccs = []
                    inter_surgeon_maes = []
                    
                    if (dataset_mode in os_vs_os_results and 
                        opmode in os_vs_os_results[dataset_mode] and 
                        angle in os_vs_os_results[dataset_mode][opmode]):
                        
                        for pair_data in os_vs_os_results[dataset_mode][opmode][angle].values():
                            inter_surgeon_icc = pair_data.get("icc", np.nan)
                            inter_surgeon_mae = pair_data.get("error_metrics", {}).get("MAE", np.nan)
                            
                            if not np.isnan(inter_surgeon_icc):
                                inter_surgeon_iccs.append(inter_surgeon_icc)
                            
                            if not np.isnan(inter_surgeon_mae):
                                inter_surgeon_maes.append(inter_surgeon_mae)
                    
                    # Calculate averages
                    avg_surgeon_icc = np.nanmean(surgeon_iccs) if surgeon_iccs else np.nan
                    avg_surgeon_mae = np.nanmean(surgeon_maes) if surgeon_maes else np.nan
                    avg_inter_surgeon_icc = np.nanmean(inter_surgeon_iccs) if inter_surgeon_iccs else np.nan
                    avg_inter_surgeon_mae = np.nanmean(inter_surgeon_maes) if inter_surgeon_maes else np.nan
                    
                    # Store the data
                    comparison_data.append({
                        "Dataset": dataset_mode,
                        "Operation Mode": opmode,
                        "Angle": angle,
                        "Mean OS vs AI ICC": mean_os_icc,
                        "Mean OS vs AI MAE": mean_os_mae,
                        "Individual OS vs AI ICC": avg_surgeon_icc,
                        "Individual OS vs AI MAE": avg_surgeon_mae,
                        "Inter-surgeon ICC": avg_inter_surgeon_icc,
                        "Inter-surgeon MAE": avg_inter_surgeon_mae
                    })
    
    if not comparison_data:
        logger.warning("No data available for comparison visualizations")
        return
    
    # Convert to DataFrame for easier plotting
    df = pd.DataFrame(comparison_data)
    
    # Create ICC comparison plot
    plt.figure(figsize=(14, 8))
    df_melted = pd.melt(df, id_vars=["Dataset", "Operation Mode", "Angle"],
                       value_vars=["Mean OS vs AI ICC", "Individual OS vs AI ICC", "Inter-surgeon ICC"],
                       var_name="Comparison", value_name="ICC")
    
    # Filter out NaN values
    df_melted = df_melted.dropna(subset=["ICC"])
    
    if len(df_melted) > 0:
        sns.boxplot(x="Angle", y="ICC", hue="Comparison", data=df_melted)
        plt.title("ICC Comparison")
        plt.xticks(rotation=45, ha="right")
        plt.grid(True, linestyle="--", alpha=0.7)
        plt.tight_layout()
        plt.savefig(compare_dir / "icc_comparison.png", dpi=300, bbox_inches="tight")
        plt.close()
    
    # Create MAE comparison plot
    plt.figure(figsize=(14, 8))
    df_melted = pd.melt(df, id_vars=["Dataset", "Operation Mode", "Angle"],
                       value_vars=["Mean OS vs AI MAE", "Individual OS vs AI MAE", "Inter-surgeon MAE"],
                       var_name="Comparison", value_name="MAE")
    
    # Filter out NaN values
    df_melted = df_melted.dropna(subset=["MAE"])
    
    if len(df_melted) > 0:
        sns.boxplot(x="Angle", y="MAE", hue="Comparison", data=df_melted)
        plt.title("MAE Comparison")
        plt.xticks(rotation=45, ha="right")
        plt.grid(True, linestyle="--", alpha=0.7)
        plt.tight_layout()
        plt.savefig(compare_dir / "mae_comparison.png", dpi=300, bbox_inches="tight")
        plt.close()
    
    # Create dataset-specific comparisons
    for dataset_mode in ["int", "ex"]:
        df_dataset = df[df["Dataset"] == dataset_mode]
        
        if len(df_dataset) == 0:
            continue
        
        # ICC by dataset mode
        plt.figure(figsize=(14, 8))
        df_melted = pd.melt(df_dataset, id_vars=["Operation Mode", "Angle"],
                          value_vars=["Mean OS vs AI ICC", "Individual OS vs AI ICC", "Inter-surgeon ICC"],
                          var_name="Comparison", value_name="ICC")
        
        # Filter out NaN values
        df_melted = df_melted.dropna(subset=["ICC"])
        
        if len(df_melted) > 0:
            sns.boxplot(x="Angle", y="ICC", hue="Comparison", data=df_melted)
            plt.title(f"ICC Comparison - {dataset_mode.upper()} Dataset")
            plt.xticks(rotation=45, ha="right")
            plt.grid(True, linestyle="--", alpha=0.7)
            plt.tight_layout()
            plt.savefig(compare_dir / f"icc_comparison_{dataset_mode}.png", dpi=300, bbox_inches="tight")
            plt.close()
        
        # MAE by dataset mode
        plt.figure(figsize=(14, 8))
        df_melted = pd.melt(df_dataset, id_vars=["Operation Mode", "Angle"],
                          value_vars=["Mean OS vs AI MAE", "Individual OS vs AI MAE", "Inter-surgeon MAE"],
                          var_name="Comparison", value_name="MAE")
        
        # Filter out NaN values
        df_melted = df_melted.dropna(subset=["MAE"])
        
        if len(df_melted) > 0:
            sns.boxplot(x="Angle", y="MAE", hue="Comparison", data=df_melted)
            plt.title(f"MAE Comparison - {dataset_mode.upper()} Dataset")
            plt.xticks(rotation=45, ha="right")
            plt.grid(True, linestyle="--", alpha=0.7)
            plt.tight_layout()
            plt.savefig(compare_dir / f"mae_comparison_{dataset_mode}.png", dpi=300, bbox_inches="tight")
            plt.close()
    
    # Create a scatter plot of ICC vs MAE
    plt.figure(figsize=(10, 8))
    
    # Prepare data for scatter plot
    scatter_data = []
    
    for row in comparison_data:
        # Mean OS vs AI
        if not np.isnan(row["Mean OS vs AI ICC"]) and not np.isnan(row["Mean OS vs AI MAE"]):
            scatter_data.append({
                "ICC": row["Mean OS vs AI ICC"],
                "MAE": row["Mean OS vs AI MAE"],
                "Type": "Mean OS vs AI",
                "Angle": row["Angle"]
            })
        
        # Individual OS vs AI
        if not np.isnan(row["Individual OS vs AI ICC"]) and not np.isnan(row["Individual OS vs AI MAE"]):
            scatter_data.append({
                "ICC": row["Individual OS vs AI ICC"],
                "MAE": row["Individual OS vs AI MAE"],
                "Type": "Individual OS vs AI",
                "Angle": row["Angle"]
            })
        
        # Inter-surgeon
        if not np.isnan(row["Inter-surgeon ICC"]) and not np.isnan(row["Inter-surgeon MAE"]):
            scatter_data.append({
                "ICC": row["Inter-surgeon ICC"],
                "MAE": row["Inter-surgeon MAE"],
                "Type": "Inter-surgeon",
                "Angle": row["Angle"]
            })
    
    if scatter_data:
        scatter_df = pd.DataFrame(scatter_data)
        
        # Create scatter plot
        sns.scatterplot(x="MAE", y="ICC", hue="Type", style="Angle", data=scatter_df, s=100)
        plt.title("Relationship between ICC and MAE")
        plt.xlabel("Mean Absolute Error (degrees)")
        plt.ylabel("Intraclass Correlation Coefficient")
        plt.grid(True, linestyle="--", alpha=0.7)
        plt.tight_layout()
        plt.savefig(compare_dir / "icc_vs_mae.png", dpi=300, bbox_inches="tight")
        plt.close()
    
    logger.info(f"Comparison visualizations saved to {compare_dir}")



def calculate_extended_error_metrics(actual: List[float], predicted: List[float]) -> Dict[str, float]:
    """
    Calculate extended error metrics between two sets of measurements.
    
    Args:
        actual (List[float]): Actual values (ground truth)
        predicted (List[float]): Predicted values
        
    Returns:
        Dict[str, float]: Dictionary of error metrics
    """
    if not actual or not predicted or len(actual) != len(predicted):
        return {metric: np.nan for metric in ERROR_METRICS + ['Within_1deg', 'Within_2deg', 'Within_3deg', 'Max_Error']}
    
    actual = np.array(actual)
    predicted = np.array(predicted)
    
    # Original metrics
    me = np.mean(predicted - actual)
    mae = np.mean(np.abs(predicted - actual))
    rmse = np.sqrt(np.mean((predicted - actual) ** 2))
    
    # Additional metrics from the second code
    # Calculate percentage within clinically acceptable thresholds
    within_1_degree = np.mean(np.abs(predicted - actual) <= 1.0) * 100
    within_2_degrees = np.mean(np.abs(predicted - actual) <= 2.0) * 100
    within_3_degrees = np.mean(np.abs(predicted - actual) <= 3.0) * 100
    
    # Maximum error
    max_error = np.max(np.abs(predicted - actual))
    
    # Median absolute error
    median_ae = np.median(np.abs(predicted - actual))
    
    return {
        "ME": me,
        "MAE": mae,
        "RMSE": rmse,
        "Median_AE": median_ae,
        "Within_1deg": within_1_degree,
        "Within_2deg": within_2_degrees,
        "Within_3deg": within_3_degrees,
        "Max_Error": max_error
    }

def alternative_calculate_icc(x: List[float], y: List[float]) -> float:
    """
    Calculate Intraclass Correlation Coefficient (ICC) using the method from the second code.
    
    Args:
        x (List[float]): First set of measurements
        y (List[float]): Second set of measurements
        
    Returns:
        float: ICC value
    """
    if not x or not y or len(x) != len(y) or len(x) < 2:
        return np.nan
    
    try:
        # Create a matrix where rows are subjects and columns are raters
        ratings = np.column_stack((x, y))
        n = len(x)
        k = 2  # Number of raters
        
        # Calculate mean for each subject
        subject_means = np.mean(ratings, axis=1)
        
        # Calculate mean for each rater
        rater_means = np.mean(ratings, axis=0)
        
        # Calculate overall mean
        overall_mean = np.mean(ratings)
        
        # Calculate between-subjects sum of squares
        ss_subjects = k * np.sum((subject_means - overall_mean) ** 2)
        
        # Calculate between-raters sum of squares
        ss_raters = n * np.sum((rater_means - overall_mean) ** 2)
        
        # Calculate total sum of squares
        ss_total = np.sum((ratings - overall_mean) ** 2)
        
        # Calculate residual sum of squares
        ss_residual = ss_total - ss_subjects - ss_raters
        
        # Calculate mean squares
        ms_subjects = ss_subjects / (n - 1)
        ms_residual = ss_residual / ((n - 1) * (k - 1))
        
        # Calculate ICC (two-way mixed, absolute agreement, single rater/measurement)
        icc = (ms_subjects - ms_residual) / (ms_subjects + (k - 1) * ms_residual)
        
        return icc
    
    except Exception as e:
        logger.warning(f"Error calculating ICC with alternative method: {e}")
        return np.nan

def perform_statistical_tests(x: List[float], y: List[float]) -> Dict[str, float]:
    """
    Perform statistical tests to compare two sets of measurements.
    
    Args:
        x (List[float]): First set of measurements
        y (List[float]): Second set of measurements
        
    Returns:
        Dict[str, float]: Dictionary of test results
    """
    if not x or not y or len(x) != len(y) or len(x) < 2:
        return {"t_test_p": np.nan, "wilcoxon_p": np.nan}
    
    try:
        # Paired t-test
        t_stat, t_p = stats.ttest_rel(x, y)
        
        # Wilcoxon signed-rank test
        w_stat, w_p = stats.wilcoxon(x, y)
        
        return {
            "t_test_p": t_p,
            "wilcoxon_p": w_p
        }
    except Exception as e:
        logger.warning(f"Error performing statistical tests: {e}")
        return {"t_test_p": np.nan, "wilcoxon_p": np.nan}

def plot_error_histogram(true_values: List[float], predicted_values: List[float], 
                        title: str, 
                        save_path: Optional[str] = None) -> plt.Figure:
    """
    Create histogram of prediction errors.
    
    Args:
        true_values (List[float]): True measurements
        predicted_values (List[float]): Predicted measurements
        title (str): Plot title
        save_path (Optional[str]): Path to save the figure
        
    Returns:
        plt.Figure: The created figure
    """
    if not true_values or not predicted_values or len(true_values) != len(predicted_values):
        logger.warning("Cannot create error histogram with empty or unequal arrays")
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.text(0.5, 0.5, "Insufficient data for error histogram", 
                horizontalalignment='center', verticalalignment='center')
        return fig
    
    # Calculate errors
    errors = np.array(predicted_values) - np.array(true_values)
    
    # Create the plot
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Plot histogram with KDE
    sns.histplot(errors, kde=True, bins=20, ax=ax)
    
    # Add vertical lines for mean and median
    ax.axvline(np.mean(errors), color='r', linestyle='--', 
              label=f'Mean: {np.mean(errors):.2f}°')
    ax.axvline(np.median(errors), color='g', linestyle='--', 
              label=f'Median: {np.median(errors):.2f}°')
    
    # Set labels and title
    ax.set_title(title)
    ax.set_xlabel('Error (Predicted - True) (degrees)')
    ax.set_ylabel('Frequency')
    ax.legend()
    
    # Add grid
    ax.grid(True, linestyle='--', alpha=0.7)
    
    # Tight layout
    plt.tight_layout()
    
    # Save if requested
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
    
    return fig

def plot_error_boxplot(all_errors: List[List[float]], angle_names: List[str], 
                     save_path: Optional[str] = None) -> plt.Figure:
    """
    Create boxplot comparing errors across all angles.
    
    Args:
        all_errors (List[List[float]]): List of error arrays for each angle
        angle_names (List[str]): Names of the angles
        save_path (Optional[str]): Path to save the figure
        
    Returns:
        plt.Figure: The created figure
    """
    if not all_errors or not angle_names or len(all_errors) != len(angle_names):
        logger.warning("Cannot create error boxplot with invalid inputs")
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.text(0.5, 0.5, "Insufficient data for error boxplot", 
                horizontalalignment='center', verticalalignment='center')
        return fig
    
    # Create the plot
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Create boxplot
    ax.boxplot(all_errors, labels=angle_names)
    
    # Set labels and title
    ax.set_title('Error Distribution Comparison')
    ax.set_ylabel('Error (Predicted - True) (degrees)')
    
    # Add grid
    ax.grid(True, axis='y', linestyle='--', alpha=0.7)
    
    # Tight layout
    plt.tight_layout()
    
    # Save if requested
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
    
    return fig

def plot_combined_error_histogram(all_errors: List[List[float]], angle_names: List[str], 
                               save_path: Optional[str] = None) -> plt.Figure:
    """
    Create combined histogram of errors for all angles.
    
    Args:
        all_errors (List[List[float]]): List of error arrays for each angle
        angle_names (List[str]): Names of the angles
        save_path (Optional[str]): Path to save the figure
        
    Returns:
        plt.Figure: The created figure
    """
    if not all_errors or not angle_names or len(all_errors) != len(angle_names):
        logger.warning("Cannot create combined error histogram with invalid inputs")
        fig, ax = plt.subplots(figsize=(12, 6))
        ax.text(0.5, 0.5, "Insufficient data for combined error histogram", 
                horizontalalignment='center', verticalalignment='center')
        return fig
    
    # Create the plot
    fig, ax = plt.subplots(figsize=(12, 6))
    
    # Create a DataFrame for easier plotting
    error_df = pd.DataFrame()
    for i, errors in enumerate(all_errors):
        error_df[angle_names[i]] = pd.Series(errors)
    
    # Plot KDE for each angle
    for angle in angle_names:
        sns.kdeplot(error_df[angle], label=angle, ax=ax)
    
    # Set labels and title
    ax.set_title('Error Distribution Comparison')
    ax.set_xlabel('Error (Predicted - True) (degrees)')
    ax.set_ylabel('Density')
    ax.legend()
    
    # Add grid
    ax.grid(True, linestyle='--', alpha=0.7)
    
    # Tight layout
    plt.tight_layout()
    
    # Save if requested
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
    
    return fig

def plot_correlation_matrix(dataset: Dict, dataset_mode: str, opmode: str, save_path: Optional[str] = None) -> plt.Figure:
    """
    Create correlation matrix for all measurements in a dataset.
    
    Args:
        dataset (Dict): Dataset containing measurements
        dataset_mode (str): Dataset mode (int/ex)
        opmode (str): Operation mode
        save_path (Optional[str]): Path to save the figure
        
    Returns:
        plt.Figure: The created figure
    """
    try:
        # Collect data for all angles
        data = {}
        display_labels = []
        
        # Extract true values for all angles (OS)
        for angle in RELEVANT_ANGLES:
            if (angle in dataset[dataset_mode][opmode]):
                values = list(dataset[dataset_mode][opmode][angle].values())
                if values:
                    data[f"{angle}_OS"] = values
                    display_labels.append(f"{angle} (OS)")
        
        # Extract predicted values for all angles (AI)
        for angle in RELEVANT_ANGLES:
            if (angle in dataset[dataset_mode][opmode]):
                values = list(dataset[dataset_mode][opmode][angle].values())
                if values:
                    data[f"{angle}_AI"] = values
                    display_labels.append(f"{angle} (AI)")
        
        if not data:
            logger.warning(f"No data available for correlation matrix: {dataset_mode}, {opmode}")
            return None
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Drop rows with NaN values
        df = df.dropna()
        
        if len(df) == 0:
            logger.warning("No valid data for correlation matrix after dropping NaNs")
            return None
        
        # Calculate correlation matrix
        corr_df = df.corr()
        
        # Create the plot
        fig, ax = plt.subplots(figsize=(12, 10))
        
        # Plot correlation matrix
        sns.heatmap(corr_df, annot=True, fmt='.2f', cmap='coolwarm', 
                   vmin=-1, vmax=1, center=0, square=True, linewidths=.5,
                   xticklabels=display_labels, yticklabels=display_labels, ax=ax)
        
        # Set title
        ax.set_title(f'Correlation Matrix - {dataset_mode.upper()}, {opmode}')
        
        # Rotate x-axis labels
        plt.xticks(rotation=45, ha='right')
        
        # Tight layout
        plt.tight_layout()
        
        # Save if requested
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
        
        return fig
    
    except Exception as e:
        logger.error(f"Error creating correlation matrix: {e}")
        return None

def plot_focused_correlation_matrix(dataset: Dict, dataset_mode: str, opmode: str, save_path: Optional[str] = None) -> plt.Figure:
    """
    Create focused correlation matrix showing only OS vs AI comparisons.
    
    Args:
        dataset (Dict): Dataset containing measurements
        dataset_mode (str): Dataset mode (int/ex)
        opmode (str): Operation mode
        save_path (Optional[str]): Path to save the figure
        
    Returns:
        plt.Figure: The created figure
    """
    try:
        # Collect data for all angles
        data = {}
        display_labels = []
        os_columns = []
        ai_columns = []
        
        # Extract true values for all angles (OS)
        for angle in RELEVANT_ANGLES:
            if (angle in dataset[dataset_mode][opmode]):
                values = list(dataset[dataset_mode][opmode][angle].values())
                if values:
                    col_name = f"{angle}_OS"
                    data[col_name] = values
                    display_labels.append(f"{angle} (OS)")
                    os_columns.append(col_name)
        
        # Extract predicted values for all angles (AI)
        for angle in RELEVANT_ANGLES:
            if (angle in dataset[dataset_mode][opmode]):
                values = list(dataset[dataset_mode][opmode][angle].values())
                if values:
                    col_name = f"{angle}_AI"
                    data[col_name] = values
                    display_labels.append(f"{angle} (AI)")
                    ai_columns.append(col_name)
        
        if not data or not os_columns or not ai_columns:
            logger.warning(f"No data available for focused correlation matrix: {dataset_mode}, {opmode}")
            return None
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Drop rows with NaN values
        df = df.dropna()
        
        if len(df) == 0:
            logger.warning("No valid data for focused correlation matrix after dropping NaNs")
            return None
        
        # Calculate correlation matrix
        corr_df = df.corr()
        
        # Create mask to keep only OS vs AI correlations
        mask = np.ones_like(corr_df, dtype=bool)
        for i, col_i in enumerate(corr_df.columns):
            for j, col_j in enumerate(corr_df.columns):
                # If one is OS and one is AI
                if ((col_i in os_columns and col_j in ai_columns) or 
                    (col_i in ai_columns and col_j in os_columns)):
                    mask[i, j] = False
        
        # Create the plot
        fig, ax = plt.subplots(figsize=(12, 10))
        
        # Plot focused correlation matrix
        cmap = sns.diverging_palette(220, 10, as_cmap=True)
        sns.heatmap(corr_df, mask=mask, annot=True, fmt='.2f', cmap=cmap,
                   vmin=-1, vmax=1, center=0, square=True, linewidths=.5,
                   xticklabels=display_labels, yticklabels=display_labels, ax=ax)
        
        # Set title
        ax.set_title(f'Focused Correlation: OS vs AI - {dataset_mode.upper()}, {opmode}')
        
        # Rotate x-axis labels
        plt.xticks(rotation=45, ha='right')
        
        # Tight layout
        plt.tight_layout()
        
        # Save if requested
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
        
        return fig
    
    except Exception as e:
        logger.error(f"Error creating focused correlation matrix: {e}")
        return None

def generate_html_report(mean_os_vs_ai_results: Dict, output_dir: Path) -> None:
    """
    Generate comprehensive HTML report for all analyses.
    
    Args:
        mean_os_vs_ai_results (Dict): Results from mean OS vs AI analysis
        individual_os_vs_ai_results (Dict): Results from individual OS vs AI analysis
        os_vs_os_results (Dict): Results from OS vs OS analysis
        output_dir (Path): Directory to save the report
    """
    html_path = output_dir / "report.html"
    
    try:
        with open(html_path, 'w') as f:
            # HTML Header
            f.write('<html><head>')
            f.write('<style>')
            f.write('body{font-family:Arial;margin:20px;line-height:1.6}')
            f.write('table{border-collapse:collapse;width:100%;margin-bottom:20px}')
            f.write('th,td{text-align:left;padding:8px;border:1px solid #ddd}')
            f.write('th{background-color:#f2f2f2}tr:nth-child(even){background-color:#f9f9f9}')
            f.write('h1,h2,h3{color:#333}')
            f.write('.flex-container{display:flex;flex-wrap:wrap;justify-content:space-between}')
            f.write('.flex-item{flex:1;min-width:300px;margin:10px}')
            f.write('</style>')
            f.write('<title>Medical Imaging Analysis Report</title></head><body>')
            f.write('<h1>Medical Imaging Analysis Report</h1>')
            
            # Summary section
            f.write('<h2>Summary</h2>')
            f.write('<p>This report presents the results of the analysis comparing measurements between AI and orthopedic surgeons.</p>')
            
            # Mean OS vs AI Results
            f.write('<h2>Mean OS vs AI Agreement</h2>')
            
            for dataset_mode in ["int", "ex"]:
                f.write(f'<h3>{dataset_mode.upper()} Dataset</h3>')
                
                for opmode in OP_MODES:
                    f.write(f'<h4>Operation Mode: {opmode}</h4>')
                    
                    # Create summary table
                    f.write('<table>')
                    f.write('<tr><th>Angle</th><th>Samples</th><th>MAE</th><th>RMSE</th><th>ICC</th><th>Within 2°</th><th>Pearson r</th></tr>')
                    
                    for angle in RELEVANT_ANGLES:
                        if (dataset_mode in mean_os_vs_ai_results and 
                            opmode in mean_os_vs_ai_results[dataset_mode] and 
                            angle in mean_os_vs_ai_results[dataset_mode][opmode]):
                            
                            data = mean_os_vs_ai_results[dataset_mode][opmode][angle]
                            
                            # Extract metrics
                            n_samples = data.get("n_samples", 0)
                            mae = data.get("error_metrics", {}).get("MAE", np.nan)
                            rmse = data.get("error_metrics", {}).get("RMSE", np.nan)
                            icc = data.get("icc", np.nan)
                            within_2deg = data.get("error_metrics", {}).get("Within_2deg", np.nan)
                            pearson_r = data.get("correlation", {}).get("pearson", np.nan)
                            
                            # Add row to table
                            f.write(f'<tr>')
                            f.write(f'<td>{angle}</td>')
                            f.write(f'<td>{n_samples}</td>')
                            f.write(f'<td>{mae:.2f if not np.isnan(mae) else "N/A"}</td>')
                            f.write(f'<td>{rmse:.2f if not np.isnan(rmse) else "N/A"}</td>')
                            f.write(f'<td>{icc:.3f if not np.isnan(icc) else "N/A"}</td>')
                            f.write(f'<td>{within_2deg:.1f}% if not np.isnan(within_2deg) else "N/A"</td>')
                            f.write(f'<td>{pearson_r:.3f if not np.isnan(pearson_r) else "N/A"}</td>')
                            f.write(f'</tr>')
                    
                    f.write('</table>')
                    
                    # Add visualizations if they exist
                    figures_dir = output_dir / "figures" / f"{dataset_mode}_{opmode}" / "mean_os_vs_ai"
                    if figures_dir.exists():
                        f.write('<div class="flex-container">')
                        
                        for angle in RELEVANT_ANGLES:
                            scatter_path = figures_dir / f"scatter_{angle}.png"
                            bland_altman_path = figures_dir / f"bland_altman_{angle}.png"
                            error_hist_path = figures_dir / f"error_histogram_{angle}.png"
                            
                            if scatter_path.exists() or bland_altman_path.exists() or error_hist_path.exists():
                                f.write(f'<h5>{angle}</h5>')
                                f.write('<div class="flex-container">')
                                
                                if scatter_path.exists():
                                    rel_path = scatter_path.relative_to(output_dir)
                                    f.write(f'<div class="flex-item"><img src="{rel_path}" style="width:100%"/></div>')
                                
                                if bland_altman_path.exists():
                                    rel_path = bland_altman_path.relative_to(output_dir)
                                    f.write(f'<div class="flex-item"><img src="{rel_path}" style="width:100%"/></div>')
                                
                                if error_hist_path.exists():
                                    rel_path = error_hist_path.relative_to(output_dir)
                                    f.write(f'<div class="flex-item"><img src="{rel_path}" style="width:100%"/></div>')
                                
                                f.write('</div>')
                        
                        f.write('</div>')
            
            # Add summary plots
            summary_dir = output_dir / "figures" / "summary"
            if summary_dir.exists():
                f.write('<h3>Summary Plots</h3>')
                f.write('<div class="flex-container">')
                
                for dataset_mode in ["int", "ex"]:
                    icc_heatmap_path = summary_dir / f"icc_heatmap_{dataset_mode}.png"
                    icc_barplot_path = summary_dir / f"icc_barplot_{dataset_mode}.png"
                    error_boxplot_path = summary_dir / f"error_boxplot_{dataset_mode}.png"
                    combined_error_hist_path = summary_dir / f"combined_error_histogram_{dataset_mode}.png"
                    
                    if any([p.exists() for p in [icc_heatmap_path, icc_barplot_path, error_boxplot_path, combined_error_hist_path]]):
                        f.write(f'<h4>{dataset_mode.upper()} Dataset Summary</h4>')
                        f.write('<div class="flex-container">')
                        
                        for path in [icc_heatmap_path, icc_barplot_path, error_boxplot_path, combined_error_hist_path]:
                            if path.exists():
                                rel_path = path.relative_to(output_dir)
                                f.write(f'<div class="flex-item"><img src="{rel_path}" style="width:100%"/></div>')
                        
                        f.write('</div>')
                
                f.write('</div>')
            
            # Finish HTML
            f.write('</body></html>')
        
        logger.info(f"HTML report generated at {html_path}")
    
    except Exception as e:
        logger.error(f"Error generating HTML report: {e}")

def analyze_mean_os_vs_ai_extended(mean_os_data: Dict, ai_data: Dict) -> Dict:
    """
    Analyze agreement between mean OS and AI measurements with extended metrics.
    
    Args:
        mean_os_data (Dict): Mean orthopedic surgeon data
        ai_data (Dict): AI data
        
    Returns:
        Dict: Analysis results
    """
    results = {}
    
    # For storing error collections for summary plots
    all_errors = {}
    
    for dataset_mode in ["int", "ex"]:
        results[dataset_mode] = {}
        all_errors[dataset_mode] = {}
        
        for op_id, opmode in enumerate(OP_MODES):
            results[dataset_mode][opmode] = {}
            all_errors[dataset_mode][opmode] = {}
            
            cur_rel_angles = RELEVANT_ANGLES_POSTOP
            if op_id == 0:
                cur_rel_angles = RELEVANT_ANGLES_PREOP
            
            for angle in RELEVANT_ANGLES:
                # Extract common measurements
                os_values, ai_values = extract_common_measurements(
                    mean_os_data, ai_data, dataset_mode, opmode, angle
                )
                
                if not os_values or not ai_values:
                    results[dataset_mode][opmode][angle] = {
                        "n_samples": 0,
                        "error_metrics": {metric: np.nan for metric in ERROR_METRICS + ['Within_1deg', 'Within_2deg', 'Within_3deg', 'Max_Error']},
                        "correlation": {method: np.nan for method in CORRELATION_METHODS},
                        "icc": np.nan,
                        "alt_icc": np.nan,
                        "statistical_tests": {"t_test_p": np.nan, "wilcoxon_p": np.nan}
                    }
                    continue
                
                # Calculate errors for summary plots
                errors = np.array(ai_values) - np.array(os_values)
                all_errors[dataset_mode][opmode][angle] = errors.tolist()
                
                # Calculate metrics
                error_metrics = calculate_extended_error_metrics(os_values, ai_values)
                correlation = calculate_correlation_metrics(os_values, ai_values)
                icc = calculate_icc(os_values, ai_values)
                alt_icc = alternative_calculate_icc(os_values, ai_values)
                statistical_tests = perform_statistical_tests(os_values, ai_values)
                
                # Store results
                results[dataset_mode][opmode][angle] = {
                    "n_samples": len(os_values),
                    "error_metrics": error_metrics,
                    "correlation": correlation,
                    "icc": icc,
                    "alt_icc": alt_icc,
                    "statistical_tests": statistical_tests
                }
                
                # Create and save plots if we have enough data
                if len(os_values) >= 3:
                    # Directory for this specific comparison
                    plot_dir = OUTPUT_DIR / "figures" / f"{dataset_mode}_{opmode}" / "mean_os_vs_ai"
                    plot_dir.mkdir(parents=True, exist_ok=True)
                    
                    # Bland-Altman plot (already in original code)
                    title = f"Bland-Altman: Mean OS vs AI - {angle} ({dataset_mode}, {opmode})"
                    save_path = plot_dir / f"bland_altman_{angle}.png"
                    create_bland_altman_plot(
                        os_values, ai_values, title=title, save_path=str(save_path)
                    )
                    
                    # Scatter plot (already in original code)
                    title = f"Correlation: Mean OS vs AI - {angle} ({dataset_mode}, {opmode})"
                    save_path = plot_dir / f"scatter_{angle}.png"
                    create_scatter_plot(
                        os_values, ai_values, title=title, 
                        xlabel="Mean OS Measurement", ylabel="AI Measurement",
                        save_path=str(save_path)
                    )
                    
                    # Error histogram plot (new)
                    title = f"Error Distribution: Mean OS vs AI - {angle} ({dataset_mode}, {opmode})"
                    save_path = plot_dir / f"error_histogram_{angle}.png"
                    plot_error_histogram(
                        os_values, ai_values, title=title, save_path=str(save_path)
                    )
            
            # Create summary plots for this dataset mode and opmode
            if all_errors[dataset_mode][opmode]:
                # Directory for summary plots
                summary_dir = OUTPUT_DIR / "figures" / "summary"
                summary_dir.mkdir(parents=True, exist_ok=True)
                
                # Extract errors and angle names for this combination
                all_error_arrays = []
                angle_names = []
                for angle in RELEVANT_ANGLES:
                    if angle in all_errors[dataset_mode][opmode]:
                        all_error_arrays.append(all_errors[dataset_mode][opmode][angle])
                        angle_names.append(angle)
                
                if all_error_arrays and angle_names:
                    # Error boxplot
                    save_path = summary_dir / f"error_boxplot_{dataset_mode}_{opmode}.png"
                    plot_error_boxplot(all_error_arrays, angle_names, save_path=str(save_path))
                    
                    # Combined error histogram
                    save_path = summary_dir / f"combined_error_histogram_{dataset_mode}_{opmode}.png"
                    plot_combined_error_histogram(all_error_arrays, angle_names, save_path=str(save_path))
                    
                    # Correlation matrices
                    save_path = summary_dir / f"correlation_matrix_{dataset_mode}_{opmode}.png"
                    combined_data = {
                        dataset_mode: {
                            opmode: {angle: {} for angle in RELEVANT_ANGLES}
                        }
                    }
                    
                    # Populate with OS data
                    for angle in RELEVANT_ANGLES:
                        if angle in mean_os_data[dataset_mode][opmode]:
                            combined_data[dataset_mode][opmode][angle] = mean_os_data[dataset_mode][opmode][angle]
                    
                    # Add AI data
                    for angle in RELEVANT_ANGLES:
                        if angle in ai_data[dataset_mode][opmode]:
                            combined_data[dataset_mode][opmode][f"{angle}_AI"] = ai_data[dataset_mode][opmode][angle]
                    
                    # Create correlation matrix
                    plot_correlation_matrix(combined_data, dataset_mode, opmode, save_path=str(save_path))
                    
                    # Create focused correlation matrix
                    save_path = summary_dir / f"focused_correlation_matrix_{dataset_mode}_{opmode}.png"
                    plot_focused_correlation_matrix(combined_data, dataset_mode, opmode, save_path=str(save_path))
    return results



def export_paired_comparison_to_excel(mean_os_data: Dict, ai_data: Dict, filename: str = "paired_comparison.xlsx") -> None:
    """
    Export paired comparison between mean OS and AI measurements to Excel.
    Creates one sheet per dataset_mode and opmode combination.
    Each sheet has groups of columns for each angle (OS, AI, and Difference), with patient ID in the first column.
    Includes conditional formatting for difference columns and summary statistics at the bottom.
    
    Args:
        mean_os_data (Dict): Mean orthopedic surgeon data
        ai_data (Dict): AI data
        filename (str): Output filename
    """
    logger.info(f"Exporting paired comparison to Excel: {filename}")
    output_path = OUTPUT_DIR / filename
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Process each dataset_mode and opmode combination
        for dataset_mode in ["int", "ex"]:
            for opmode in OP_MODES:
                # Create sheet name from dataset_mode and opmode
                sheet_name = f"{dataset_mode}_{opmode}"
                logger.info(f"Creating sheet: {sheet_name}")
                
                # Get all patient IDs that have at least one measurement in either dataset
                all_patient_ids = set()
                
                # Check mean OS data
                for angle in RELEVANT_ANGLES:
                    try:
                        if (angle in mean_os_data[dataset_mode][opmode]):
                            all_patient_ids.update(mean_os_data[dataset_mode][opmode][angle].keys())
                    except KeyError:
                        continue
                
                # Check AI data
                for angle in RELEVANT_ANGLES:
                    try:
                        if (angle in ai_data[dataset_mode][opmode]):
                            all_patient_ids.update(ai_data[dataset_mode][opmode][angle].keys())
                    except KeyError:
                        continue
                
                if not all_patient_ids:
                    logger.warning(f"No patient IDs found for {sheet_name}, skipping sheet")
                    continue
                
                # Create data dictionary for this sheet
                sheet_data = {"Patient_ID": sorted(all_patient_ids)}
                
                # Add groups of columns for each angle (OS, AI, Diff)
                for angle in RELEVANT_ANGLES:
                    # Initialize columns for OS, AI, and Difference
                    os_column = []
                    ai_column = []
                    diff_column = []
                    
                    # Populate data for each patient
                    for patient_id in sheet_data["Patient_ID"]:
                        # Get OS value
                        os_value = None
                        try:
                            if (angle in mean_os_data[dataset_mode][opmode] and 
                                patient_id in mean_os_data[dataset_mode][opmode][angle]):
                                os_value = mean_os_data[dataset_mode][opmode][angle][patient_id]
                        except KeyError:
                            pass
                        
                        # Get AI value
                        ai_value = None
                        try:
                            if (angle in ai_data[dataset_mode][opmode] and 
                                patient_id in ai_data[dataset_mode][opmode][angle]):
                                ai_value = ai_data[dataset_mode][opmode][angle][patient_id]
                        except KeyError:
                            pass
                        
                        # Calculate difference if both values exist and round to two decimal places
                        diff_value = None
                        if os_value is not None and ai_value is not None:
                            # Round OS and AI values to two decimal places
                            os_value = round(float(os_value), 2)
                            ai_value = round(float(ai_value), 2)
                            diff_value = round(abs(os_value - ai_value), 2)
                        
                        # Add values to columns
                        os_column.append(os_value)
                        ai_column.append(ai_value)
                        diff_column.append(diff_value)
                    
                    # Add columns to data dictionary
                    sheet_data[f"{angle}_OS"] = os_column
                    sheet_data[f"{angle}_AI"] = ai_column
                    sheet_data[f"{angle}_Diff"] = diff_column
                
                # Convert to DataFrame
                df = pd.DataFrame(sheet_data)
                
                # Create a new DataFrame with the ordered columns
                final_df = pd.DataFrame(index=df.index)
                final_df["Patient_ID"] = df["Patient_ID"]
                
                # List to store all difference columns for summary statistics
                diff_columns = []
                
                for angle in RELEVANT_ANGLES:
                    if f"{angle}_OS" in df.columns:
                        final_df[f"{angle}_OS"] = df[f"{angle}_OS"]
                    else:
                        final_df[f"{angle}_OS"] = None
                    
                    if f"{angle}_AI" in df.columns:
                        final_df[f"{angle}_AI"] = df[f"{angle}_AI"]
                    else:
                        final_df[f"{angle}_AI"] = None
                    
                    if f"{angle}_Diff" in df.columns:
                        final_df[f"{angle}_Diff"] = df[f"{angle}_Diff"]
                        diff_columns.append(f"{angle}_Diff")
                    else:
                        final_df[f"{angle}_Diff"] = None
                
                # Calculate summary statistics for difference columns
                summary_row = {"Patient_ID": "Summary"}
                for col in final_df.columns:
                    if col == "Patient_ID":
                        continue
                    
                    if col.endswith("_Diff"):
                        # Calculate mean and std for difference columns (properly excluding None, NaN and empty values)
                        # Use pandas for more robust calculation
                        values = pd.Series([v for v in final_df[col] if v is not None and pd.notna(v) and v != ""])
                        if not values.empty:
                            mean_value = values.mean()
                            std_value = values.std()
                            summary_row[col] = f"Mean: {mean_value:.2f}, Std: {std_value:.2f}"
                        else:
                            summary_row[col] = "No data"
                    else:
                        summary_row[col] = ""
                
                # Write to Excel
                final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the sheet for formatting
                worksheet = writer.sheets[sheet_name]
                
                # Add summary row at the bottom
                summary_row_idx = len(final_df) + 2  # +2 because Excel is 1-indexed and we have a header row
                for col_idx, column in enumerate(final_df.columns, start=1):
                    cell = worksheet.cell(row=summary_row_idx, column=col_idx)
                    cell.value = summary_row[column]
                    
                    # Make summary row bold
                    from openpyxl.styles import Font
                    cell.font = Font(bold=True)
                
                # Apply conditional formatting to difference columns
                from openpyxl.styles import PatternFill
                from openpyxl.formatting.rule import ColorScaleRule
                
                for col_idx, column in enumerate(final_df.columns, start=1):
                    if column.endswith("_Diff"):
                        # Get column letter
                        col_letter = get_column_letter(col_idx)
                        
                        # Apply color scale (white to red)
                        # The range includes the data rows but excludes the header and summary row
                        data_range = f"{col_letter}2:{col_letter}{len(final_df) + 1}"
                        
                        color_scale_rule = ColorScaleRule(
                            start_type='min', start_color='FFFFFF',  # White for low values
                            end_type='max', end_color='FF0000'      # Red for high values
                        )
                        
                        worksheet.conditional_formatting.add(data_range, color_scale_rule)
                
                # Auto-adjust column widths
                for i, column in enumerate(final_df.columns):
                    column_width = max(
                        final_df[column].astype(str).map(len).max(),
                        len(column)
                    ) + 2
                    
                    # Set reasonable width limits
                    column_width = min(max(column_width, 10), 30)  # Min width 10, max width 30
                    
                    # Convert to Excel column letter and set width
                    try:
                        worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width
                    except Exception as e:
                        logger.warning(f"Couldn't set column width for column {i + 1}: {e}")
    
    logger.info(f"Paired comparison exported to {output_path}")



def main():
    """
    Main function to run the entire analysis pipeline.
    """
    logger.info("Starting medical imaging analysis")
    
    # Load data
    logger.info("Loading surgeon data")
    os_data, unique_ids = load_surgeon_data()
    
    logger.info("Loading AI data")
    ai_data = load_ai_data(unique_ids)
    
    # Calculate mean OS data
    logger.info("Calculating mean orthopedic surgeon measurements")
    mean_os_data = calculate_mean_os_data(os_data)
    
    # Export paired comparison to Excel (new function call)
    logger.info("Exporting paired comparison to Excel")
    export_paired_comparison_to_excel(mean_os_data, ai_data)
    
    # Run analyses with extended metrics
    logger.info("Analyzing mean OS vs AI agreement with extended metrics")
    mean_os_vs_ai_results = analyze_mean_os_vs_ai_extended(mean_os_data, ai_data)
    
    logger.info("Analyzing individual OS vs AI agreement")
    #individual_os_vs_ai_results = analyze_individual_os_vs_ai(os_data, ai_data)
    
    logger.info("Analyzing inter-observer agreement between surgeons")
    #os_vs_os_results = analyze_os_vs_os(os_data)
    
    # Export results
    logger.info("Exporting results to CSV")
    export_results_to_csv(mean_os_vs_ai_results, "mean_os_vs_ai_results.csv")
    #export_results_to_csv(individual_os_vs_ai_results, "individual_os_vs_ai_results.csv")
    #export_results_to_csv(os_vs_os_results, "os_vs_os_results.csv")
    
    # Create summary tables
    logger.info("Creating summary tables")
    #create_summary_tables(mean_os_vs_ai_results, individual_os_vs_ai_results, os_vs_os_results)
    
    # Export all metrics to Excel
    logger.info("Exporting all metrics to Excel")
    #export_all_metrics_to_excel(mean_os_vs_ai_results, individual_os_vs_ai_results, os_vs_os_results)
    
    # Create summary plots
    logger.info("Creating summary plots")
    create_summary_plots(mean_os_vs_ai_results)
    
    # Generate HTML report
    logger.info("Generating HTML report")
    generate_html_report(mean_os_vs_ai_results, OUTPUT_DIR)
    
    logger.info("Analysis complete!")
    return mean_os_vs_ai_results
                
if __name__ == "__main__":
    mean_os_vs_ai_results = main()
